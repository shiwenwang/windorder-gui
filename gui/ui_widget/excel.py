
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, colors, Alignment


class ExcelAPI:

    def __init__(self, filename):
        self.filename = filename
        self.read()

    def read(self):
        self.wb = load_workbook(self.filename)
        self.wb.create_sheet('Tower Info')

    def update(self, wind_info, tower_info):
        for sheet in self.wb:
            self._update_sheet(sheet, wind_info, tower_info)

    def write(self, save_filename):
        self.wb.save(save_filename)
                    
    def _update_sheet(self, sheet, wind_info, tower_info):
        first_col = self.wb['Site Condition']['A']
        for cell in first_col:
            if not cell.internal_value:
                max_row = cell.row - 3
                break
        try:
            sites_count = max_row
        except UnboundLocalError:
            sites_count = len(first_col) - 2

        if sheet.title == 'Tower Info':
            _map = {'塔架编号': '', '塔架受限': 'tower_limit', '极限比例': 'ul_prop',
                    '疲劳比例': 'fl_prop', '风载属性': 'wind_limit', '塔架重量(t)': 'weight', '标准规范': 'std_spec'}
            for c, name in enumerate(_map.keys()):
                _cell = sheet.cell(1, c + 1, name)
                _cell.alignment = Alignment(horizontal='center', vertical='center')
                _cell.font = Font(name='微软雅黑', size=10)
            for r in range(len(tower_info)):
                tower_id = list(tower_info.keys())[r]
                tower = tower_info[tower_id]
                for c, name in enumerate(_map.keys()):
                    if 0 == c:
                        value = tower_id
                    else:
                        value = tower[_map[name]]
                    _cell = sheet.cell(r + 2, c + 1, value)
                    _cell.alignment = Alignment(horizontal='center', vertical='center')
                    _cell.font = Font(name='微软雅黑', size=10)
        elif sheet.title == 'Site Condition':
            labels = list(tower_info.keys())
            for r, label in enumerate(labels):

                new_row_id = sites_count + 4 + r
                sheet.cell(row=new_row_id, column=1, value=label)
                sheet.cell(row=new_row_id, column=1).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=new_row_id, column=1).font = \
                    Font(name=sheet.cell(3, 2).font.name, size=sheet.cell(3, 2).font.size, color=colors.DARKBLUE)
                sheet.cell(row=new_row_id, column=7, value=wind_info[label]['condition']['ρ'][label])
                sheet.cell(row=new_row_id, column=7).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=new_row_id, column=7).font = \
                    Font(name=sheet.cell(3, 2).font.name, size=sheet.cell(3, 2).font.size, color=colors.DARKBLUE)
                sheet.cell(row=new_row_id, column=8, value=wind_info[label]['condition']['vave'][label])
                sheet.cell(row=new_row_id, column=8).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=new_row_id, column=8).font = \
                    Font(name=sheet.cell(3, 2).font.name, size=sheet.cell(3, 2).font.size, color=colors.DARKBLUE)
                sheet.cell(row=new_row_id, column=9, value=wind_info[label]['condition']['a'][label])
                sheet.cell(row=new_row_id, column=9).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=new_row_id, column=9).font = \
                    Font(name=sheet.cell(3, 2).font.name, size=sheet.cell(3, 2).font.size, color=colors.DARKBLUE)
                sheet.cell(row=new_row_id, column=10, value=wind_info[label]['condition']['k'][label])
                sheet.cell(row=new_row_id, column=10).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=new_row_id, column=10).font = \
                    Font(name=sheet.cell(3, 2).font.name, size=sheet.cell(3, 2).font.size, color=colors.DARKBLUE)
                sheet.cell(row=new_row_id, column=13, value=wind_info[label]['condition']['α'][label])
                sheet.cell(row=new_row_id, column=13).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=new_row_id, column=13).font = \
                    Font(name=sheet.cell(3, 2).font.name, size=sheet.cell(3, 2).font.size, color=colors.DARKBLUE)
                sheet.cell(row=new_row_id, column=14, value=wind_info[label]['condition']['θmean'][label])
                sheet.cell(row=new_row_id, column=14).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=new_row_id, column=14).font = \
                    Font(name=sheet.cell(3, 2).font.name, size=sheet.cell(3, 2).font.size, color=colors.DARKBLUE)
                sheet.cell(row=new_row_id, column=15, value=wind_info[label]['condition']['v50'][label])
                sheet.cell(row=new_row_id, column=15).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=new_row_id, column=15).font = \
                    Font(name=sheet.cell(3, 2).font.name, size=sheet.cell(3, 2).font.size, color=colors.DARKBLUE)
        elif sheet.title in ['M=1', 'M=10', 'ETM']:
            map = {'M=1': 'm1', 'M=10': 'm10', 'ETM': 'etm'}

            for cr in sheet.merged_cells.ranges.copy():
                sheet.unmerge_cells(cr.coord)

            sheet.delete_cols(sites_count + 3, 4)
            
            labels = list(tower_info.keys())
            for c, label in enumerate(labels):
                new_col_id = sites_count + 3 + c                                
                sheet.cell(row=1, column=new_col_id, value=label)
                sheet.cell(row=1, column=new_col_id).font = \
                    Font(name=sheet.cell(1, 2).font.name, size=sheet.cell(1, 2).font.size, color=colors.DARKBLUE)
                sheet.cell(row=1, column=new_col_id).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=2, column=new_col_id, value=sheet.cell(2, 2).value)
                sheet.cell(row=2, column=new_col_id).font = \
                    Font(name=sheet.cell(2, 2).font.name, size=sheet.cell(2, 2).font.size, color=colors.DARKBLUE)
                sheet.cell(row=2, column=new_col_id).alignment = Alignment(horizontal='center', vertical='center')
                for i, ti in enumerate(wind_info[label][map[sheet.title]][label]):
                    sheet.cell(row=i + 3, column=new_col_id, value=ti)
                    sheet.cell(i + 3, new_col_id).alignment = Alignment(horizontal='center', vertical='center')
                    sheet.cell(i + 3, new_col_id).font = \
                        Font(name=sheet.cell(i + 3, 2).font.name, size=sheet.cell(i + 3, 2).font.size, color=colors.DARKBLUE)
        else:
            pass


if __name__ == "__main__":

    filepath = r'E:\WorkSpace\6_Programming\wind_order\wind_order-gui\files\wind\shunfeng_140-2.5-140m_341.5t.xlsx'
    filepath1 = r'E:\WorkSpace\6_Programming\wind_order\wind_order-gui\files\wind\shunfeng_140-2.5-140m_341.5t_save.xlsx'
    excel = Excel(filepath,filepath1)
    excel.read()
    